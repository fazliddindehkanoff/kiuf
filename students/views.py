from io import BytesIO

from django.db import transaction
from django.http import FileResponse
from django.shortcuts import render
from django.contrib.auth.decorators import login_required

from .models import Student
from .utils import generate_pdf, generate_qr_code_titles
from openpyxl import load_workbook


@login_required(login_url="login")
def home_view(request):
    if request.method == "POST" and request.FILES:
        file = request.FILES['file']

        student_ids = []

        wb = load_workbook(file, read_only=True)
        sheet_names = wb.sheetnames

        with transaction.atomic():
            for sheet_name in sheet_names:
                ws = wb[sheet_name]

                for row in ws.iter_rows(min_row=2, values_only=True):
                    first_name = row[1] or "-"
                    last_name = row[0] or "-"
                    middle_name = row[2] or "-"
                    passport_number = row[3] or "-"
                    birth_date = row[4] or "-"
                    otm = row[5] or "-"
                    speciality = row[6] or "-"
                    study_type = row[7] or "-"
                    edu_type = row[8] or "-"
                    document_number = row[9] or "-"
                    registered_date = row[10] or "-"

                    student = Student(
                        first_name=first_name,
                        last_name=last_name,
                        middle_name=middle_name,
                        passport_number=passport_number,
                        birth_date=birth_date,
                        otm=otm,
                        speciality=speciality,
                        study_type=study_type,
                        edu_type=edu_type,
                        document_number=document_number,
                        registered_date=registered_date
                    )
                    student.save()
                    student_ids.append(student.id)

        qr_code_titles = generate_qr_code_titles(ids=student_ids)
        pdf_bytes = generate_pdf(qr_code_titles)

        response = FileResponse(BytesIO(pdf_bytes))
        response['Content-Type'] = 'application/pdf'
        response['Content-Disposition'] = 'attachment; filename="results.pdf"'

        return response

    return render(request, "home.html")


def students_view(request, pk):
    student = Student.objects.filter(pk=pk).first()
    if student:
        return render(request, "student_detail.html", {"student": student})
    return render(request, "404.html")
